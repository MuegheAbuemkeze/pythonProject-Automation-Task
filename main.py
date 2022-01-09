# Python script to update the emails of all employees stored in an employee database in
# Excel and csv file formats from something@helpinghands.cm
# to something@helpinghands.org


#importing all neccessary libraries
from openpyxl import Workbook, load_workbook
import pandas as pd

# Load workbook from desktop
wb = load_workbook('/home/mac/Desktop/employeedata.xlsx')
ws = wb.active  # Gives active worksheet of open workbook

# Storing the updated email in a dictionary
Email_Updates = {
    "something@helpinghands.cm": "something@helpinghands.org"
}

# Loop through the rows and update their values by accessing the updated email stored in the 
#dictionary called Email_Updates 
for rowNum in range(1, 31):
    Updates = ws.cell(row=rowNum, column=2).value
    if Updates in Email_Updates:
        ws.cell(row=rowNum, column=2).value = Email_Updates[Updates]


wb.save('/home/mac/Desktop/employeedata.xlsx')



#Reading the csv file
df = pd.read_csv("/home/mac/Desktop/employeedata.csv")

#updating the column data
df['Email address'] = df['Email address'].replace({'something@helpinghands.cm':'something@helpinghands.org'})

#writing into the csv file
df.to_csv("/home/mac/Desktop/employeedata.csv", index=False) 
